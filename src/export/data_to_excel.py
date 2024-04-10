import openpyxl
import pandas as pd
from openpyxl import Workbook

import openpyxl
from openpyxl import Workbook

def append_pdf_to_excel(company_name, website, pdf_link, filename):
    filename = f"./data/output/{filename}.xlsx"
    try:
        try:
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Company Name", "Website", "PDF Link"])

        if sheet.max_row == 1 and sheet['A1'].value is None:
            sheet.append(["Company Name", "Website", "PDF Link"])

        sheet.append([company_name, website, pdf_link])
        workbook.save(filename)
    except Exception as e:
        print(f"Failed to append to Excel: {e}")
