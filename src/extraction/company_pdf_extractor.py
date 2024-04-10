from openpyxl import load_workbook

def extract_ei_pairs_from_excel(e_column='E', i_column='I'):
    # wb = load_workbook(file_path)
    wb = load_workbook("/home/zeynep/Projects/side-projects/llm-project/data/input/ai_automation_project.xlsx")
    sheet = wb.active

    # Convert column letters to numbers (1-indexed)
    e_column_number = ord(e_column.upper()) - 64
    i_column_number = ord(i_column.upper()) - 64

    # Initialize an empty list to store pairs
    ei_pairs = []

    # Iterate through each row, starting from row 2 to skip the header
    for i in range(2, sheet.max_row + 1):
        # Fetch values from the E-th and I-th columns
        e_value = sheet.cell(row=i, column=e_column_number).value
        i_value = sheet.cell(row=i, column=i_column_number).value

        # Check if both cells are not empty
        if e_value is not None and i_value is not None:
            # Add the pair to the list
            ei_pairs.append((e_value, i_value))

    return ei_pairs

if __name__ == "__main__":
    extract_ei_pairs_from_excel()