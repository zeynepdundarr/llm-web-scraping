from openpyxl import load_workbook

def extract_company_names_from_excel(file_path, column='E'):

    wb = load_workbook(file_path)
    sheet = wb.active
    
    column_number = ord(column.upper()) - 64
    
    company_names_set = {sheet.cell(row=i, column=column_number).value for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=column_number).value is not None}
    
    company_names_list = list(company_names_set)
    # extract 2 companies for test purposes
    return company_names_list


if __name__ == "__main__":
    print(extract_company_names_from_excel("./data/input/ai_automation_project.xlsx"))