from openpyxl import load_workbook

def extract_company_names_from_excel(file_path, column='E'):
    # Load the workbook and select the active sheet
    wb = load_workbook(file_path)
    sheet = wb.active  # or wb['SheetName'] if you know the sheet name
    
    # Determine the column number from the column letter
    column_number = ord(column.upper()) - 64  # Convert column letter to column number
    
    # Extract company names from the specified column
    company_names = [sheet.cell(row=i, column=column_number).value for i in range(1, sheet.max_row + 1) if sheet.cell(row=i, column=column_number).value is not None]
    
    return company_names